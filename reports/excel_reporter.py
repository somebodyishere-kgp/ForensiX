"""
Excel Report Generator for ForensiX
Converts metadata results into formatted Excel workbooks
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, 
    Side
)
from openpyxl.utils import get_column_letter


def generate_excel_report(results, output_path):
    """
    Generate comprehensive Excel report from collected metadata
    
    Parameters:
    -----------
    results : list
        List of metadata dictionaries from collect_metadata()
    output_path : str
        Where to save the Excel file (e.g., "output/report.xlsx")
    
    Example:
    --------
    results = [
        {"filename": "image.jpg", "camera_make": "Canon", ...},
        {"filename": "document.pdf", "author": "John", ...}
    ]
    generate_excel_report(results, "output/forensix_report.xlsx")
    """
    
    # Ensure output directory exists
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Sheet 1: Summary (Overview)
        summary_df = _create_summary_sheet(results)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Detailed Metadata
        detailed_df = _create_detailed_sheet(results)
        detailed_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Sheet 3: Suspicious Findings
        suspicious_df = _create_suspicious_sheet(results)
        if not suspicious_df.empty:
            suspicious_df.to_excel(writer, sheet_name='Suspicious', index=False)
        
        # Sheet 4: GPS Data (if any)
        gps_df = _create_gps_sheet(results)
        if not gps_df.empty:
            gps_df.to_excel(writer, sheet_name='GPS Data', index=False)
    
    # Format the workbook
    _format_workbook(output_path)
    
    print(f"✓ Excel report generated: {output_path}")


def _create_summary_sheet(results):
    """
    Summary sheet: One row per file with key info
    """
    summary_data = []
    
    for result in results:
        summary_data.append({
            "Filename": result.get("filename", "Unknown"),
            "File Size": result.get("file_size", "N/A"),
            "MIME Type": result.get("mime_type", "N/A"),
            "Modified Date": result.get("modified", "N/A"),
            "MD5": result.get("md5", "N/A")[:16] + "..." if result.get("md5") else "N/A",
            "Suspicious Findings": len(result.get("suspicious_findings", [])),
            "Error": result.get("error", "OK") if result.get("error") else "OK"
        })
    
    return pd.DataFrame(summary_data)


def _create_detailed_sheet(results):
    """
    Detailed sheet: All metadata fields (except raw_metadata)
    """
    detailed_data = []
    
    for result in results:
        row = {}
        
        for key, value in result.items():
            # Skip raw_metadata as it's too large for Excel
            if key == "raw_metadata":
                continue
            
            # Skip suspicious findings (has own sheet)
            if key == "suspicious_findings":
                continue
            
            # Convert lists to comma-separated strings
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            
            row[key] = value
        
        detailed_data.append(row)
    
    df = pd.DataFrame(detailed_data)
    
    # Ensure Filename is first column
    cols = df.columns.tolist()
    if "filename" in cols:
        cols.remove("filename")
        cols = ["filename"] + cols
        df = df[cols]
    
    return df


def _create_suspicious_sheet(results):
    """
    Suspicious sheet: All red flags found
    """
    suspicious_data = []
    
    for result in results:
        filename = result.get("filename", "Unknown")
        findings = result.get("suspicious_findings", [])
        
        for finding in findings:
            suspicious_data.append({
                "Filename": filename,
                "Severity": finding.get("severity", "INFO"),
                "Finding": finding.get("finding", "Unknown"),
                "Evidence": finding.get("evidence", "N/A")
            })
    
    return pd.DataFrame(suspicious_data)


def _create_gps_sheet(results):
    """
    GPS sheet: Location data from photos
    """
    gps_data = []
    
    for result in results:
        latitude = result.get("gps_latitude")
        longitude = result.get("gps_longitude")
        
        if latitude and longitude:
            gps_data.append({
                "Filename": result.get("filename", "Unknown"),
                "Latitude": latitude,
                "Longitude": longitude,
                "Timestamp": result.get("created", "N/A")
            })
    
    return pd.DataFrame(gps_data)


def _format_workbook(output_path):
    """
    Apply formatting to the Excel workbook
    """
    wb = load_workbook(output_path)
    
    # Define colors
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )
    
    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=11
    )
    
    error_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )
    
    warning_fill = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                 min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Highlight errors in red
                if "error" in str(cell.value).lower() and cell.value != "OK":
                    cell.fill = error_fill
                
                # Highlight HIGH severity in orange
                if cell.value == "HIGH":
                    cell.fill = warning_fill
        
        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    wb.save(output_path)